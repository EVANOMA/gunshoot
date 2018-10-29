from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

wb = Workbook()
ws = wb.active

# 对齐
align = Alignment(horizontal='center')
# 字体样式
Price_ft = Font(size=20, bold=True)
Title_ft = Font(size=15, bold=True)
Character_ft = Font(size=12)
# 背景填充
# 颜色选择：http://www.114la.com/other/rgb.htm
Price_fill = PatternFill("solid", fgColor="FFC0CB")
Title_fill = PatternFill("solid", fgColor="FFA500")

Character_fill = PatternFill("solid", fgColor="F0F0F0")
# 自我长度不一致
Error1_fill = PatternFill("solid", fgColor="EE9572")
# 表之间长度不一致
Error2_fill = PatternFill("solid", fgColor="EEE0E5")
# 物品ID或数量不一致
Error3_fill = PatternFill("solid", fgColor="CD0000")
Fill = {
            0: Character_fill,
            1: Error1_fill,
            2: Error2_fill,
            3: Error3_fill,
        }


def headline(row, col, price, s):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    # 赋值
    d = ws.cell(row=row, column=col, value=price)
    d.font = Price_ft
    d.alignment = align
    d.fill = Price_fill

    row = row + 1
    title = \
        {
            0: "period_recharge_reward_table",
            1: "period_recharge_sum_table",
            2: s,
        }
    j = 0
    for i in range(1, 9, 3):
        ws.merge_cells(start_row=row, start_column=i, end_row=row, end_column=i + 2)
        d = ws.cell(row=row, column=i, value=title[j])
        d.font = Title_ft
        d.alignment = align
        d.fill = Title_fill
        j = j + 1
    wb.save('test.xlsx')
    return row


def write(fill_num, l, row, col):
    for x in l:
        d = ws.cell(row=row, column=col, value=x)
        d.font = Character_ft
        d.alignment = align
        d.fill = Fill[fill_num]
        row = row + 1
    wb.save('test.xlsx')


def write_s(fill_num, s, row, col):
    d = ws.cell(row=row, column=col, value=s)
    d.font = Character_ft
    d.alignment = align
    d.fill = Fill[fill_num]
    wb.save('test.xlsx')


def mark(row, col):
    for i in range(col, col+7, 3):
        d = ws.cell(row=row, column=i)
        d.fill = Fill[3]
    wb.save('test.xlsx')


def adjustment():
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
    wb.save('test.xlsx')
